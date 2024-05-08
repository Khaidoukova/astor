from django.conf import settings
from django.contrib.auth.mixins import (LoginRequiredMixin,
                                        PermissionRequiredMixin,
                                        UserPassesTestMixin)
from django.core.mail import send_mail
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Border, Side
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.utils import timezone

from django.views.generic import (ListView, DetailView,
                                  CreateView, UpdateView,
                                  DeleteView, TemplateView, View)
from openpyxl.styles import numbers
from announcements.models import Announcement
from news.models import News
from .forms import CompanyForm, OfficeForm, User_requestForm, BookingForm, CarsForm
from .models import Company, Office, User_request, Booking, Cars
from .tasks import request_to_telegram


def contacts(request):
    return render(request, 'main/contacts.html')


class IndexView(TemplateView):
    template_name = 'main/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['announcements'] = Announcement.objects.all().order_by('-created_at')
        context['news'] = News.objects.all()

        return context


class CompanyListView(ListView):
    model = Company

    def get_queryset(self):
        companies = Company.objects.all()
        return companies


class CompanyDetailView(DetailView):
    model = Company


class CompanyCreateView(CreateView):
    model = Company
    form_class = CompanyForm
    success_url = reverse_lazy('main:index')


class CompanyUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    success_url = reverse_lazy('main:index')

    def test_func(self):
        user = self.request.user

        if user.is_staff or user.is_superuser or user.is_manager:
            return True
        else:
            return False


class CompanyDeleteView(LoginRequiredMixin,
                        DeleteView):
    model = Company
    success_url = reverse_lazy('main:index')


class OfficeListView(ListView):
    model = Office


class OfficeDetailView(DetailView):
    model = Office


class OfficeCreateView(CreateView):
    model = Office
    form_class = OfficeForm
    success_url = reverse_lazy('main:index')


class OfficeUpdateView(LoginRequiredMixin,
                       UserPassesTestMixin,
                       UpdateView):
    model = Office
    form_class = OfficeForm
    success_url = reverse_lazy('main:index')

    def test_func(self):
        user = self.request.user

        if user.is_staff or user.is_superuser or user.is_manager:
            return True
        else:
            return False


class OfficeDeleteView(LoginRequiredMixin,
                       PermissionRequiredMixin,
                       DeleteView):
    model = Office
    success_url = reverse_lazy('main:index')


class User_requestListView(ListView):
    model = User_request
    template_name = 'main/user_request_list.html'


    def get_queryset(self):
        queryset = User_request.objects.all()
        active_user_requests = queryset.filter(status=User_request.IN_PROGRESS)
        done_user_requests = queryset.filter(status=User_request.DONE)
        closed_user_requests = queryset.filter(status=User_request.CLOSED)
        return {
            'active_user_requests': active_user_requests,
            'done_user_requests': done_user_requests,
            'closed_user_requests': closed_user_requests,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = [status for status in User_request.STATUS if status[0] != User_request.CLOSED]
        queryset = self.get_queryset()
        context.update(queryset)
        return context


class User_requestStatusUpdateView(View):
    def post(self, request, *args, **kwargs):
        request_id = kwargs['pk']
        new_status = request.POST.get('new_status')

        req = get_object_or_404(User_request, pk=request_id)

        if new_status != req.status:  # Проверка изменения статуса
            req.status = new_status

            if new_status == User_request.DONE:  # Проверка статуса на "выполнено"
                req.date_completed = timezone.now()  # Запись текущей даты выполнения
            elif req.date_completed is not None:  # Проверка наличия даты выполнения
                req.date_completed = None  # Сброс даты выполнения, если статус изменен на другой

            req.save()
        return redirect('main:request_list')



class User_requestDetailView(DetailView):
    model = User_request


class User_requestCreateView(LoginRequiredMixin, CreateView):
    model = User_request
    form_class = User_requestForm
    success_url = reverse_lazy('main:index')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial'] = {'owner': self.request.user}
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.owner = self.request.user
        self.object = form.save()
        if self.object.urgency == 'срочно':  # Check if urgency is "срочно"
            request_to_telegram(self.object.description, self.object.office_id)  # Call the function to send telegram message
        return super().form_valid(form)


class User_requestUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = User_request
    form_class = User_requestForm
    success_url = reverse_lazy('main:index')

    def test_func(self):
        user = self.request.user

        if user.is_staff or user.is_superuser or user.is_manager:
            return True
        else:
            return False

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs


class User_requestDeleteView(LoginRequiredMixin,
                             PermissionRequiredMixin,
                             DeleteView):
    model = User_request
    success_url = reverse_lazy('main:index')
    # permission_required = 'main.delete_testcategory'


def send_feedback(request):
    IN_PROGRESS = User_request.IN_PROGRESS
    if request.method == 'POST':
        request_id = request.POST.get('request_id')
        feedback = request.POST.get('feedback')

        # Найти заявку по ID
        req = User_request.objects.get(id=request_id)

        # Добавить замечания к заявке
        req.feedback = feedback
        # Изменить статус заявки на "в работе"
        req.status = IN_PROGRESS
        req.save()

        return redirect(reverse('users:user_detail', kwargs={'pk': req.owner.pk}))

    return render(request, 'users/user_detail.html')


class BookingCreateView(CreateView):
    model = Booking
    form_class = BookingForm
    success_url = reverse_lazy('main:index')

    def form_valid(self, form):
        self.object = form.save()
        self.object.owner = self.request.user
        self.object.save()
        recipients = ['olga@noran.ru', 'khaidoukova@inbox.ru']
        subject = f'Новое бронирование переговорки от {self.object.owner}'
        message = f'Новое бронирование переговорки от {self.object.owner}:\nДата: {self.object.date}\nДлительность: {self.object.get_duration_display()}'
        from_email = settings.EMAIL_HOST_USER
        send_mail(subject, message, from_email, recipients)

        return super().form_valid(form)


class CarsListView(ListView):
    model = Cars
    template_name = 'main/car_list.html'

    def get_queryset(self):
        queryset = Cars.objects.all()
        approved_cars = queryset.filter(status=Cars.APPROVED)
        in_progress_cars = queryset.filter(status=Cars.IN_PROGRESS)
        expired_cars = queryset.filter(end_date__lt=timezone.now())
        return {
            'approved_cars': approved_cars,
            'in_progress_cars': in_progress_cars,
            'expired_cars': expired_cars,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context.update(queryset)
        return context

    def render_to_response(self, context, **response_kwargs):
        if 'export_approved' in self.request.GET:
            # Создаем новый файл Excel
            wb = openpyxl.Workbook()
            ws = wb.active

            # Добавляем заголовки столбцов
            headers = ['№', 'Номер авто', 'Модель', 'Период', 'Начало доступа', 'Окончание доступа', 'Арендатор',
                       'Офис арендатора']
            ws.append(headers)

            # Применяем стиль границ для ячеек
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # Автоматически устанавливаем ширину колонок для заголовков
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].auto_fit = True

            # Уменьшаем ширину колонки с номером вдвое
            ws.column_dimensions[get_column_letter(1)].width = ws.column_dimensions[get_column_letter(1)].width / 2

            # Добавляем данные из approved_cars
            approved_cars = context['approved_cars']
            row_number = 1
            for car in approved_cars:
                row = [row_number, car.car_id, car.model, car.period, car.start_date, car.end_date, car.owner.phone,
                       car.owner.office_number]
                ws.append(row)

                # Применяем стиль границ для строки
                for cell in ws[ws._current_row]:
                    cell.border = thin_border

                # Устанавливаем формат ячеек для дат
                ws.cell(row=ws._current_row, column=5).number_format = 'dd.mm.yy'
                ws.cell(row=ws._current_row, column=6).number_format = 'dd.mm.yy'

                row_number += 1

            # Автоматически устанавливаем ширину колонок после добавления данных
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].auto_fit = True

            # Устанавливаем параметр для печати всех данных по ширине на одну страницу
            ws.print_area = 'A1:%s%d' % (get_column_letter(ws.max_column), ws.max_row)
            ws.fit_to_pages = 1, 0  # Ширина, высота

            # Создаем HTTP-ответ для скачивания файла
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="approved_cars.xlsx"'
            wb.save(response)
            return response
        else:
            return super().render_to_response(context, **response_kwargs)

class CarsDetailView(DetailView):
    model = Cars


class CarsCreateView(CreateView):
    model = Cars
    template_name = 'main/cars_form.html'
    form_class = CarsForm
    success_url = reverse_lazy('main:index')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        self.object = form.save()
        self.object.owner = self.request.user
        self.object.save()

        return super().form_valid(form)


class CarsUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Cars
    form_class = CarsForm
    success_url = reverse_lazy('main:cars_list')

    def test_func(self):
        user = self.request.user

        if user.is_staff or user.is_superuser or user.is_manager:
            return True
        else:
            return False

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs


class CarsDeleteView(LoginRequiredMixin, DeleteView):
    model = Cars
    success_url = reverse_lazy('main:index')



def info(request):
    return render(request, 'main/info.html')
